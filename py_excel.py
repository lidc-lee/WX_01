# coding=utf-8

"""
@version: ??
@author: AA-ldc
@license: Apache Licence 
@file: py_excel.py
@time: 2016/12/15 14:27
处理Excel电子表格
"""
import openpyxl
import os

# 打开excel文档，返回一个workbook数据类型的值
wb = openpyxl.load_workbook("example.xlsx")
# # 获取工作簿的所有表名的列表
# wb.get_sheet_names()
# sheet = wb.get_sheet_by_name('Sheet1')
# active = wb.get_active_sheet()
# # 从表中取得单元格
# # 取得Cell对象
# c = sheet['A1']
# # 取得Cell对象的value的值
# title = c.value
# type(wb)
# # 写入excel文档
# # 创建一个空的Workbook对象
# wb1 = openpyxl.Workbook()
# sheet1 = wb1.get_active_sheet()
# sheet1.title = 'Hello world!!!'
# # 保存excel文档
# sheet = wb.get_active_sheet()
# sheet.title = "LDC"
# wb.save('example.xlsx')

# 创建和删除工作表
# wb.create_sheet('LEE')
# wb.remove_sheet(wb.get_sheet_by_name('LEE1'))
# wb.save('example_copy.xlsx')
# 将值写入单元格
sheet = wb.get_sheet_by_name('LDC')
sheet['A1'] = 'Hello world!!'
wb.save('example.xlsx')
print sheet['A1'].value
